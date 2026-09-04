#!/usr/bin/env python3
"""Generate a content calendar workbook from a JSON plan.

Usage: python3 build_calendar.py plan.json --out calendar.xlsx
Plan format: see ../assets/example-plan.json
"""
import argparse
import json
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DEFAULT_DAY_ORDER = ["Tue", "Thu", "Sat", "Mon", "Wed", "Fri", "Sun"]
STATUSES = ["Idea", "Drafted", "In review", "Approved", "Scheduled", "Published"]
COLUMNS = [
    "ID", "Publish date", "Day", "Time", "Platform", "Format", "Pillar", "Campaign / tentpole",
    "Hook / headline", "Caption", "Visual / video description", "Asset link", "CTA and link",
    "Hashtags / keywords", "Owner", "Status", "Paid amplification", "Paid budget", "KPI target", "Result", "Notes",
]
REACTIVE = "Reactive / trend"
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
TENTPOLE_FILL = PatternFill("solid", fgColor="DBEAFE")
REACTIVE_FILL = PatternFill("solid", fgColor="F1F5F9")


class Balancer:
    """Assigns categories so the running mix tracks target shares (largest deficit first)."""

    def __init__(self, shares):
        self.shares = shares
        self.counts = Counter()
        self.total = 0

    def next(self):
        self.total += 1
        best = max(self.shares, key=lambda k: self.shares[k] * self.total - self.counts[k])
        self.counts[best] += 1
        return best


def parse_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date()


def posting_days(channel):
    days = channel.get("days")
    if days:
        return [d for d in WEEKDAYS if d in days]
    n = max(1, min(7, int(round(channel.get("posts_per_week", 3)))))
    return [d for d in WEEKDAYS if d in DEFAULT_DAY_ORDER[:n]]


def daterange(start, end):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def build_rows(plan):
    start, end = parse_date(plan["start"]), parse_date(plan["end"])
    pillars = {p["name"]: p["share"] for p in plan["pillars"]}
    reactive_share = plan.get("reactive_share", 0.15)
    owner = plan.get("default_owner", "")
    tentpoles = plan.get("tentpoles", [])
    tentpole_index = {}
    for t in tentpoles:
        for ch in t.get("channels", []):
            tentpole_index[(ch, parse_date(t["date"]))] = t

    rows = []
    for channel in plan["channels"]:
        name = channel["name"]
        code = channel.get("code") or "".join(w[0] for w in name.split()).upper()
        days = set(posting_days(channel))
        pillar_balancer = Balancer(pillars)
        format_balancer = Balancer(channel.get("formats", {"Post": 1.0}))
        reactive_every = int(round(1 / reactive_share)) if reactive_share else 0
        slot_counter = 0
        week_counter = defaultdict(int)
        seen_tentpoles = set()

        for day in daterange(start, end):
            weekday = WEEKDAYS[day.weekday()]
            tentpole = tentpole_index.get((name, day))
            if weekday not in days and not tentpole:
                continue
            iso_year, iso_week, _ = day.isocalendar()
            week_key = f"{iso_year}-W{iso_week:02d}"
            week_counter[week_key] += 1
            slot_id = f"{code}-{week_key}-{week_counter[week_key]}"

            if tentpole:
                seen_tentpoles.add(id(tentpole))
                pillar = tentpole.get("pillar") or pillar_balancer.next()
                if pillar in pillar_balancer.shares:
                    pillar_balancer.counts[pillar] += 1
                    pillar_balancer.total += 1
                fmt = tentpole.get("format") or format_balancer.next()
                rows.append(dict(
                    id=slot_id, date=day, day=weekday, time=channel.get("time", ""), platform=name,
                    format=fmt, pillar=pillar, campaign=tentpole["name"], hook=tentpole.get("hook", ""),
                    owner=owner, status="Idea", kind="tentpole",
                ))
                continue

            slot_counter += 1
            if reactive_every and slot_counter % reactive_every == 0:
                rows.append(dict(
                    id=slot_id, date=day, day=weekday, time=channel.get("time", ""), platform=name,
                    format=format_balancer.next(), pillar=REACTIVE, campaign="",
                    hook="Hold for trend or reactive content; fill in the week of publishing.",
                    owner=owner, status="Idea", kind="reactive",
                ))
                continue

            rows.append(dict(
                id=slot_id, date=day, day=weekday, time=channel.get("time", ""), platform=name,
                format=format_balancer.next(), pillar=pillar_balancer.next(), campaign="", hook="",
                owner=owner, status="Idea", kind="regular",
            ))
    rows.sort(key=lambda r: (r["date"], r["platform"], r["id"]))
    return rows


def style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def write_workbook(plan, rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendar"
    ws.append(COLUMNS)
    style_header(ws)
    for r in rows:
        ws.append([
            r["id"], r["date"], r["day"], r["time"], r["platform"], r["format"], r["pillar"], r["campaign"],
            r["hook"], "", "", "", "", "", r["owner"], r["status"], "No", None, "", "", "",
        ])
        ws.cell(row=ws.max_row, column=2).number_format = "yyyy-mm-dd"
        fill = TENTPOLE_FILL if r["kind"] == "tentpole" else REACTIVE_FILL if r["kind"] == "reactive" else None
        if fill:
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    last = ws.max_row
    status_dv = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True)
    paid_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(paid_dv)
    status_dv.add(f"P2:P{max(last, 2)}")
    paid_dv.add(f"Q2:Q{max(last, 2)}")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(last, 2)}"
    widths = {"A": 16, "B": 12, "C": 6, "D": 7, "E": 11, "F": 10, "G": 20, "H": 26, "I": 44, "J": 60,
              "K": 40, "L": 18, "M": 24, "N": 24, "O": 14, "P": 12, "Q": 10, "R": 10, "S": 16, "T": 16, "U": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=last):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Summary
    ss = wb.create_sheet("Summary")
    ss.append([f"{plan.get('client', 'Client')} content calendar, {plan.get('period_label', '')}"])
    ss["A1"].font = Font(bold=True, size=14)
    ss.append([f"{parse_date(plan['start'])} to {parse_date(plan['end'])}; {len(rows)} slots; reactive hold {int(plan.get('reactive_share', 0.15) * 100)}%"])
    ss.append([])
    months = sorted({(r["date"].year, r["date"].month) for r in rows})
    month_labels = [date(y, m, 1).strftime("%b %Y") for y, m in months]
    ss.append(["Posts per channel per month"] + month_labels + ["Total"])
    header_row = ss.max_row
    for channel in plan["channels"]:
        counts = [sum(1 for r in rows if r["platform"] == channel["name"] and (r["date"].year, r["date"].month) == ym) for ym in months]
        ss.append([channel["name"]] + counts + [sum(counts)])
    ss.append([])
    ss.append(["Pillar mix (excluding tentpoles and reactive)", "Target", *[f"{c['name']} actual" for c in plan["channels"]]])
    for p in plan["pillars"]:
        row = [p["name"], p["share"]]
        for channel in plan["channels"]:
            regular = [r for r in rows if r["platform"] == channel["name"] and r["kind"] == "regular"]
            row.append(round(sum(1 for r in regular if r["pillar"] == p["name"]) / len(regular), 3) if regular else None)
        ss.append(row)
        for col in range(2, 3 + len(plan["channels"])):
            ss.cell(row=ss.max_row, column=col).number_format = "0%"
    ss.append([])
    ss.append(["Format mix per channel", "Format", "Target", "Actual"])
    for channel in plan["channels"]:
        ch_rows = [r for r in rows if r["platform"] == channel["name"]]
        for fmt, share in channel.get("formats", {"Post": 1.0}).items():
            actual = round(sum(1 for r in ch_rows if r["format"] == fmt) / len(ch_rows), 3) if ch_rows else None
            ss.append([channel["name"], fmt, share, actual])
            ss.cell(row=ss.max_row, column=3).number_format = "0%"
            ss.cell(row=ss.max_row, column=4).number_format = "0%"
    for cell in ss[header_row]:
        cell.font = Font(bold=True)
    ss.column_dimensions["A"].width = 44
    for col in "BCDEFG":
        ss.column_dimensions[col].width = 16

    # Tentpoles
    ts = wb.create_sheet("Tentpoles")
    ts.append(["Date", "Name", "Channels", "Format", "Pillar", "Hook"])
    style_header(ts)
    for t in sorted(plan.get("tentpoles", []), key=lambda t: t["date"]):
        ts.append([parse_date(t["date"]), t["name"], ", ".join(t.get("channels", [])), t.get("format", ""), t.get("pillar", ""), t.get("hook", "")])
        ts.cell(row=ts.max_row, column=1).number_format = "yyyy-mm-dd"
    for col, width in zip("ABCDEF", (12, 34, 24, 12, 22, 60)):
        ts.column_dimensions[col].width = width

    # Guide
    gs = wb.create_sheet("Guide")
    gs.append(["Column", "Meaning"])
    style_header(gs)
    meanings = {
        "ID": "Stable reference: channel code, ISO week, slot number in week.",
        "Publish date": "Planned publish date.",
        "Day": "Weekday, for cadence checks.",
        "Time": "Planned posting time in the plan's timezone: " + plan.get("timezone", "not stated") + ".",
        "Platform": "Channel the post publishes on.",
        "Format": "Native format for the platform (Reel, Carousel, Video, Static, Story, Text).",
        "Pillar": "Content pillar. 'Reactive / trend' rows are held open and filled the week of publishing.",
        "Campaign / tentpole": "Launch, sale or moment the post belongs to, if any.",
        "Hook / headline": "First line or first frame. Must stand alone.",
        "Caption": "Full caption in brand voice, one call to action.",
        "Visual / video description": "What the creative shows, shot list or design notes.",
        "Asset link": "Link to the final asset in the asset library.",
        "CTA and link": "The single call to action and its destination.",
        "Hashtags / keywords": "Platform-appropriate tags and search keywords.",
        "Owner": "Person responsible for getting the post to Published.",
        "Status": " > ".join(STATUSES) + ".",
        "Paid amplification": "Yes if the post will be boosted or used in ads.",
        "Paid budget": "Planned spend for this post.",
        "KPI target": "What this post is meant to achieve, in numbers.",
        "Result": "Actual result after publishing.",
        "Notes": "Anything else the team needs to know.",
    }
    for col in COLUMNS:
        gs.append([col, meanings.get(col, "")])
    gs.column_dimensions["A"].width = 26
    gs.column_dimensions["B"].width = 90

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("plan_json")
    parser.add_argument("--out", default="content-calendar.xlsx")
    args = parser.parse_args()
    with open(args.plan_json) as fh:
        plan = json.load(fh)
    for key in ("start", "end", "channels", "pillars"):
        if key not in plan:
            sys.exit(f"plan is missing required key: {key}")
    if abs(sum(p["share"] for p in plan["pillars"]) - 1) > 0.01:
        sys.exit("pillar shares must sum to 1")
    rows = build_rows(plan)
    write_workbook(plan, rows, args.out)
    by_platform = Counter(r["platform"] for r in rows)
    kinds = Counter(r["kind"] for r in rows)
    print(f"{len(rows)} slots: " + ", ".join(f"{k} {v}" for k, v in by_platform.items()))
    print(f"tentpoles {kinds['tentpole']}, reactive holds {kinds['reactive']}, regular {kinds['regular']}")
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
