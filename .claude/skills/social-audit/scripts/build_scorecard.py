#!/usr/bin/env python3
"""Build a social audit scorecard workbook from an audit JSON file.

Usage: python3 build_scorecard.py audit.json --out scorecard.xlsx
Schema: see ../references/scoring-rubric.md
"""
import argparse
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

DIMENSIONS = [
    ("foundations", "Profile and foundations", 0.10),
    ("strategy", "Content strategy and pillars", 0.15),
    ("craft", "Format mix and craft", 0.15),
    ("cadence", "Cadence and consistency", 0.10),
    ("engagement", "Engagement and community", 0.15),
    ("growth", "Growth and reach", 0.10),
    ("conversion", "Conversion and journey", 0.15),
    ("measurement", "Measurement and operations", 0.10),
]
BANDS = [(4.2, "Leading"), (3.4, "Solid"), (2.6, "Developing"), (0, "At risk")]
FILLS = {1: "F8D7DA", 2: "FBE5D6", 3: "FFF3CD", 4: "D9EAD3", 5: "B7E1CD"}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")


def band(score):
    for threshold, label in BANDS:
        if score >= threshold:
            return label
    return "At risk"


def weighted(scores):
    total, weight_sum = 0.0, 0.0
    for key, _, weight in DIMENSIONS:
        if key in scores:
            total += scores[key] * weight
            weight_sum += weight
    return round(total / weight_sum, 2) if weight_sum else None


def rate(numerator, denominator):
    if numerator is None or not denominator:
        return None
    return round(numerator / denominator, 4)


def style_header(ws, row=1, freeze=True):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if freeze:
        ws.freeze_panes = f"A{row + 1}"


def autosize(ws, minimum=10, maximum=60):
    for column in ws.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max(minimum, min(maximum, width + 2))


def build(data, out_path):
    wb = Workbook()
    channels = data.get("channels", [])
    weeks = data.get("period_weeks")

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"{data.get('client', 'Client')} social audit scorecard"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {data.get('period', '')}    Data basis: {data.get('data_basis', 'not stated')}"])
    ws.append([])
    header = ["Dimension", "Weight"] + [c["name"] for c in channels]
    ws.append(header)
    header_row = ws.max_row
    for key, label, weight in DIMENSIONS:
        row = [label, weight]
        for c in channels:
            row.append(c.get("scores", {}).get(key))
        ws.append(row)
        for idx, c in enumerate(channels, start=3):
            cell = ws.cell(row=ws.max_row, column=idx)
            if isinstance(cell.value, int) and cell.value in FILLS:
                cell.fill = PatternFill("solid", fgColor=FILLS[cell.value])
                cell.alignment = Alignment(horizontal="center")
        ws.cell(row=ws.max_row, column=2).number_format = "0%"
    overall = [weighted(c.get("scores", {})) for c in channels]
    ws.append(["Overall (weighted, out of 5)", ""] + overall)
    ws.append(["Overall (%)", ""] + [round(s / 5, 3) if s is not None else None for s in overall])
    for idx in range(3, 3 + len(channels)):
        ws.cell(row=ws.max_row, column=idx).number_format = "0%"
    ws.append(["Grade band", ""] + [band(s) if s is not None else "" for s in overall])
    for r in range(ws.max_row - 2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
    style_header(ws, header_row, freeze=False)
    ws.append([])
    ws.append(["Evidence"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Channel", "Dimension", "Evidence"])
    style_header(ws, ws.max_row, freeze=False)
    for c in channels:
        for key, label, _ in DIMENSIONS:
            note = c.get("evidence", {}).get(key)
            if note:
                ws.append([c["name"], label, note])
    autosize(ws)

    # Metrics sheet
    ms = wb.create_sheet("Metrics")
    ms.append([
        "Channel", "Handle", "Followers", "Follower growth %", "Posts", "Posts per week",
        "Reach", "Impressions", "Engagements", "Engagement rate", "Rate basis",
        "Save rate", "Share rate", "Avg engagements per post", "Video views", "Avg response (hours)",
    ])
    style_header(ms)
    for c in channels:
        followers = c.get("followers")
        start = c.get("followers_start")
        posts = c.get("posts")
        engagements = sum(v for v in (c.get("likes"), c.get("comments"), c.get("saves"), c.get("shares")) if v)
        engagements = engagements or None
        reach = c.get("reach")
        per_post = engagements / posts if engagements and posts else None
        # By reach: total engagements over total reach. By followers: average engagements per post over followers.
        if reach:
            basis, eng_rate = "by reach", rate(engagements, reach)
            save_rate, share_rate = rate(c.get("saves"), reach), rate(c.get("shares"), reach)
        else:
            basis, eng_rate = ("by followers" if followers else ""), rate(per_post, followers)
            save_rate = rate(c.get("saves") / posts, followers) if c.get("saves") and posts else None
            share_rate = rate(c.get("shares") / posts, followers) if c.get("shares") and posts else None
        growth = rate(followers - start, start) if followers is not None and start else None
        ms.append([
            c.get("name"), c.get("handle"), followers, growth, posts,
            round(posts / weeks, 1) if posts and weeks else None,
            reach, c.get("impressions"), engagements,
            eng_rate, basis, save_rate, share_rate,
            round(per_post, 1) if per_post else None,
            c.get("video_views"), c.get("avg_response_hours"),
        ])
        r = ms.max_row
        for col in (4, 10, 12, 13):
            ms.cell(row=r, column=col).number_format = "0.00%"
    autosize(ms)

    # Competitors sheet
    cs = wb.create_sheet("Competitors")
    cs.append(["Name", "Channel", "Followers", "Posts", "Posts per week", "Avg engagements per post", "Engagement rate by followers", "Notes"])
    style_header(cs)
    for c in channels:
        followers = c.get("followers")
        posts = c.get("posts")
        engagements = sum(v for v in (c.get("likes"), c.get("comments"), c.get("saves"), c.get("shares")) if v) or None
        per_post = round(engagements / posts, 1) if engagements and posts else None
        cs.append([
            f"{data.get('client', 'Client')} (client)", c.get("name"), followers, posts,
            round(posts / weeks, 1) if posts and weeks else None, per_post,
            rate(per_post, followers), "Client baseline",
        ])
        cs.cell(row=cs.max_row, column=7).number_format = "0.00%"
        cs.cell(row=cs.max_row, column=1).font = Font(bold=True)
    for comp in data.get("competitors", []):
        posts = comp.get("posts")
        per_post = comp.get("avg_engagement_per_post")
        cs.append([
            comp.get("name"), comp.get("channel"), comp.get("followers"), posts,
            round(posts / weeks, 1) if posts and weeks else None, per_post,
            rate(per_post, comp.get("followers")), comp.get("notes"),
        ])
        cs.cell(row=cs.max_row, column=7).number_format = "0.00%"
    autosize(cs)

    # Rubric sheet
    rs = wb.create_sheet("Rubric")
    rs.append(["Key", "Dimension", "Weight"])
    style_header(rs)
    for key, label, weight in DIMENSIONS:
        rs.append([key, label, weight])
        rs.cell(row=rs.max_row, column=3).number_format = "0%"
    rs.append([])
    rs.append(["Grade bands", "Leading 4.2+", "Solid 3.4 to 4.1", "Developing 2.6 to 3.3", "At risk below 2.6"])
    rs.append(["Scores are 1 to 5 integers; 3 is adequate; 5 needs best-in-class evidence. See references/scoring-rubric.md for anchors."])
    autosize(rs)

    wb.save(out_path)
    return overall


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("audit_json")
    parser.add_argument("--out", default="social-audit-scorecard.xlsx")
    args = parser.parse_args()
    with open(args.audit_json) as fh:
        data = json.load(fh)
    overall = build(data, args.out)
    for c, s in zip(data.get("channels", []), overall):
        print(f"{c['name']}: {s} / 5 ({band(s) if s is not None else 'no scores'})")
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
